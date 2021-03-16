# -*- coding: utf-8 -*-
"""
Created on Tue Mar 16 06:25:50 2021

@author: Rohan Roy
"""

path ="P:/8th SEM/INTERNSHIP/Store Management/trial_store/LnT.xlsx"
from openpyxl import load_workbook
wb = load_workbook(path)
ws=wb.get_sheet_by_name('Sheet1')
s=ws.max_row# variable to store max rows for sl num
maxr=ws.max_row
data=[]
name=input("Enter Name: ")
for i in range(1,s+1):
    if(ws.cell(row=i,column=1).value == name):
        for j in range (1,5):
            print(ws.cell(row=i,column=j).value)
            data.append(ws.cell(row=i,column=j).value)
print(data)


ws=wb.get_sheet_by_name('Sheet2')
s=ws.max_row# variable to store max rows for sl num
maxr=ws.max_row
for i in range(1,s+1):
    if(ws.cell(row=i,column=1).value == name):
        print(ws.cell(row=i,column=4).value)
        data.append(ws.cell(row=i,column=4).value)
print(data)




#print(wb.sheetnames)

if 'Sheet0' not in wb.sheetnames:
    ws = wb.create_sheet('Sheet0')
    print("CREATING")
    s=ws.max_row# variable to store max rows for sl num

    for i in range(1,6):
        ws.cell(row=s+1,column=i).value= data[i-1]
    wb.save(path)
