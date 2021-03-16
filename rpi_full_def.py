import PySimpleGUI as sg
import time
import json
import os
import sys
import PIL
import numpy as np    
import openpyxl
import qrcode
from datetime import datetime
from openpyxl import load_workbook
from xlrd import open_workbook
from openpyxl.styles import Font

#"/home/pi/Store Management/trial_store/
tqr=['Sl.No', 'Date & Time', 'ITEM PL NO:', 'Description', 'Quantity', 'Unit', 'Cost', 'Remarks']
path_ch="/home/pi/Store Management/trial_store/check.xlsx"
path="/home/pi/Store Management/trial_store/test2.xlsx"
path_ns ="/home/pi/Store Management/trial_store/NS_check.xlsx"


#------------------------------------minimum stock   START--------------------------------------
def min_check():
    path_stock ="/home/pi/Store Management/trial_store/check.xlsx"
    #path_stock=p
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    wb=load_workbook(path_stock)
    cpl=[]
    t_minstk=[]
    t_minstk=["PL Number:","Minimum Stock:","Description:","Existing Stock:"]
    cpl_l=[]
    ws=wb.get_sheet_by_name("PL")
    s=ws.max_row
    for i in range (1,s+1):
        if(ws.cell(row=i,column=4).value)<(ws.cell(row=i,column=2).value):
            #print((ws.cell(row=i,column=4).value),(ws.cell(row=i,column=2).value))
            #item_stock=ws.cell(row=i,column=2).value
            for j in range(1,5):
                cpl_l.append(t_minstk[j-1]+str((ws.cell(row=i,column=j).value)))
            #ws.cell(row=i,column=4).value=t
            
            cpl.append(cpl_l)

            cpl_l=[]
            #print("Des stck",des_stk)
        
        
    ws=wb.get_sheet_by_name("PL2")
    
    s=ws.max_row
    
    
    for i in range (1,s+1):
        #print("IN NS")
        if(ws.cell(row=i,column=4).value)<(ws.cell(row=i,column=2).value):   
            #item_stock=ws.cell(row=i,column=2).value
            for j in range(1,5):
                cpl_l.append(t_minstk[j-1]+str((ws.cell(row=i,column=j).value)))
            
          
            cpl.append(cpl_l)
            cpl_l=[]
          
            
            
            #print("PL NUMBER:",p,"Total Quantity: ",item_tot)   
    wb.save(path_stock) 
    rr=""
    for dta in cpl:
        rr=rr+str(dta)+"\n\n"
    sg.Popup("UNDER STOCK ITEMS\n",rr) 
    
    
    
    
#------------------------------------minimum stock  END-----------------------------------------
 




#------------------------------------UPDATE TOTAL START--------------------------------------
def update_Tot(p1,t1):
    path_stock ="/home/pi/Store Management/trial_store/check.xlsx"
    #path_stock=p
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    wb=load_workbook(path_stock)
    p=p1
    cpl=[]
    item_stock=0
    cpl_l=[]
    ws=wb.get_sheet_by_name("PL")
    t=t1
    s=ws.max_row
    for i in range (1,s+1):
        if(ws.cell(row=i,column=1).value==p):   
            #item_stock=ws.cell(row=i,column=2).value
            
            ws.cell(row=i,column=4).value=t
            
            #cpl.append(cpl_l)
            #print("Des stck",des_stk)
        
        
    ws=wb.get_sheet_by_name("PL2")
    
    s=ws.max_row
    for i in range (1,s+1):
        if(ws.cell(row=i,column=1).value==p):   
            #item_stock=ws.cell(row=i,column=2).value
            
            ws.cell(row=i,column=4).value=t
            
            #cpl.append(cpl_l)
            #print("Des stck",des_stk)
            
          
            
            
          
            
            
            #print("PL NUMBER:",p,"Total Quantity: ",item_tot)   
    wb.save(path_stock)     
    
    
    
    
#------------------------------------UPDATE TOTAL END-----------------------------------------





#-------------------------------------SEARCH START-------------------------------------



def search():
    path1="/home/pi/Store Management/trial_store/test2.xlsx"
    path2="/home/pi/Store Management/trial_store/ledgers_dmtr.xlsx"
    
    
    
    i_item=[]
    i2_item=[]#HEADINGS OF ROW
    a_item=[]
    count=0
    
    layout_search = [
             [sg.Text('PL. '), sg.In(key=3)],[sg.Button('Capture')],
                    
            [sg.Frame(layout=[
                    #[sg.Checkbox('Checkbox', size=(10,1)),  sg.Checkbox('My second checkbox!', default=True)],
                    [sg.Radio('IN', "RADIO1",), sg.Radio('OUT', "RADIO1",), sg.Radio('DMTR', "RADIO1",) ]], title='Select the file  ',title_color='red', relief=sg.RELIEF_SUNKEN, tooltip='Select the mode')],
            [sg.Submit()]
            ]
    
    
        
    window3=sg.Window('Incoming Items',layout_search ,default_element_size=(40, 1), grab_anywhere=False)
    event, values = window3.read()
    
    sg.Popup('Title',
         'The results of the window.',
         'The button clicked was "{}"'.format(event),
         'The values are', values)
    if values[0]==True:
        book = open_workbook(path1)
        sheet=book.sheet_by_index(0)
        title_count=10
    if values[1]==True:
        book = open_workbook(path1)
        sheet=book.sheet_by_index(1)
        title_count=11
    if values[2]==True:
        book = open_workbook(path2)
        sheet=book.sheet_by_index(0)
        title_count=10
        
        
    
    
    
    #d=['Sl. No.','P.L.NO.','DESCRIPTION','UNIT','AAC','Page No.','SOH AC','SOH MSD','Consm  19-20' ]
    if (event=='Capture'):
        
        ic=capture()
    
        values[3]=ic
        
    i=values[3]            
    if (values[3]==None or values[3]==""):
        sg.Popup('RECORD NOT EXISTING')
        window3.close()
    elif (values[3]!=None and values[3]!=""):   
        print("\n\n")
        row1=sheet.row(0)
        for rowidx in range(sheet.nrows):
            row = sheet.row(rowidx)
    
            for (colidx, cell) in enumerate(row):
                if cell.value ==i:
                    count=count+1
                    print("-----------------------------------")
                    print("Result",count)
                    print(sheet.name)
                    r=rowidx
                    c=colidx
                    print("row,col=",r,",",c)
                    """txt=sheet.cell(2,2).value# command to access cell value
                    print(txt)
                    val="over written"
                    sheet.cell(2,2).value=val
                    print("Val:",sheet.cell(2,2).value)"""
                    for j in range(0,title_count):
                        print(row1[j].value,":",row[j].value)
                        #ret=(row1[j].value+str(':'),row[j].value)
                        #ret=(row[j].value)
                        ret=(row1[j].value+str(':')+str(row[j].value))
                        i_item.append(ret)
                        #print("RET",)
                    #sg.Popup("Record Found")
                    
                    a_item.append(i_item)
                    i_item=[]
                    print("-----------------------------------")
        #a_item.append(i2_item)
        
        if count!=0:
            length=len(a_item)
            print("Length=",length)
            sg.Popup(count,"RESULTS FOUND")      
            rr=""
            for dta in a_item[(length-10):]:
               
               rr=rr+str(dta)+"\n\n" 
            sg.Popup("SEARCH RESULTS\n",rr)
                
        if count==0:
            #print("Record not Found")
            sg.Popup("Record not Found")

    window3.close()
#------------------------------------SEARCH END--------------------------------------
        



#-----------------------------------CHECK QUANTITY START-----------------------------

def check_Tot():
    path_tot ="/home/pi/Store Management/trial_store/ledgers_dmtr.xlsx"
    from openpyxl import load_workbook
    #from openpyxl.styles import Font
    wb=load_workbook(path_tot)
    
    layout_cap=[
                    [sg.Text('PL. '), sg.In(key=3)],[sg.Button('Capture')],
                    [sg.Submit()]
            ]
    windowcap=sg.Window('Capture PL', layout_cap,default_element_size=(40, 1), grab_anywhere=False)
    event, values1 = windowcap.read()
    
    
    

    if event=='Capture':
        print("CAMERA ")
        pl=capture()
        values1[3]=pl
    p=values1[3]
    sg.Popup(
         'PL NUMBER:', values1[3])
    windowcap.Close()
        
    print(" PL: ",p)
    for sheet in wb.sheetnames:
        
        if sheet==p:
            
            ws=wb.get_sheet_by_name(p)
            s=ws.max_row
              
            item_tot=ws.cell(row=s,column=9).value
            item_des=ws.cell(row=1,column=1).value
            
            print(item_des,"\nTotal Quantity: ",item_tot)
            print("\n\n")
            
            sg.Popup(item_des,"\nCurrent Stock: ",item_tot )
    
    wb.save(path_tot)



#----------------------------------CHECK QUANTITY END----------------------------------

















#---------------------------------------CHECK TOT1 START  (CHECKS CURRENT STOCK IN LEDGERS) ----------------------------
def check_Tot1(T):
    path_tot ="/home/pi/Store Management/trial_store/ledgers_dmtr.xlsx"
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    wb=load_workbook(path_tot)
    p=T
    for sheet in wb.sheetnames:
        
        if sheet==p:
            
            ws=wb.get_sheet_by_name(p)
            s=ws.max_row
          
            item_tot=ws.cell(row=s,column=9).value
            
            
            print("PL NUMBER:",p,"Total Quantity: ",item_tot)
            print("\n\n")
            return(item_tot)
    wb.save(path_tot)
    
        
#---------------------------------------CHECK TOT1 END----------------------------


#------------------------------------CHECK STOCK START--------------------------------------
def check_Stock(p1):
    path_stock ="/home/pi/Store Management/trial_store/check.xlsx"
    #path_stock=p
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    wb=load_workbook(path_stock)
    p=p1
    cpl=[]
    item_stock=0
    cpl_l=[]
    ws=wb.get_sheet_by_name("PL")
    
    s=ws.max_row
    for i in range (1,s+1):
        if(ws.cell(row=i,column=1).value==p):   
            item_stock=ws.cell(row=i,column=2).value
            
            des_stk=ws.cell(row=i,column=3).value
            cpl=[item_stock,des_stk]
            #cpl.append(cpl_l)
            #print("Des stck",des_stk)
            return(cpl)
        
    ws=wb.get_sheet_by_name("PL2")
    
    s=ws.max_row
    for i in range (1,s+1):
        if(ws.cell(row=i,column=1).value==p):   
            item_stock=ws.cell(row=i,column=2).value
            
            des_stk=ws.cell(row=i,column=3).value
            cpl=[item_stock,des_stk]
            #cpl.append(cpl_l)
            #print("Des stck",des_stk)
            return(cpl)
          
            
            
          
            
            
            #print("PL NUMBER:",p,"Total Quantity: ",item_tot)   
    wb.save(path_stock)     
    
    
    
    
#------------------------------------CHECK STOCK END-----------------------------------------




#-----------------------------------CHECK START--------------------------------------------

def check_pl(p_c):
    path_ch="/home/pi/Store Management/trial_store/check.xlsx"
    import openpyxl
    from xlrd import open_workbook
    from openpyxl import load_workbook
    book = open_workbook(path_ch)
    count=0
    #d=['Sl. No.','P.L.NO.','DESCRIPTION','UNIT','AAC','Page No.','SOH AC','SOH MSD','Consm  19-20' ]
    
    i=p_c
    print("\n\n")
    for sheet in book.sheets():
        for rowidx in range(sheet.nrows):
            row = sheet.row(rowidx)

            for (colidx, cell) in enumerate(row):
                if cell.value ==i:
                    count=count+1
                    print("Record found")
                    print("-----------------------------------")
                    print("Result",count)
                    print(sheet.name)
                    r=rowidx
                    c=colidx
                    print("row,col=",r,",",c)
                    
                    
                   
                    
                    
                    for j in range(0,0):
                        print(row[j].value)
            
                    print("-----------------------------------")
                    
                    return count
            
    if count==0:
            print("Record not Found")
            
            return count
            
        



#----------------------------------CHECK END------------------------------------------------
            
        



#--------------------------------------------CAMERA START----------------------------------------------
def get_image():
  with picamera.PiCamera() as camera:
    camera.resolution = (1024, 768)
    camera.start_preview()
    # Camera warm-up time
    time.sleep(5)
    camera.capture('foo.jpg')
    



def capture():

    print ('Taking picture..')
    try:
        f = 1
        #qr_count = len(os.listdir('qr_codes'))
        get_image()
        #os.system('sudo fswebcam -d /dev/video'+sys.argv[1]+' -q qr_codes/qr_'+str(qr_count)+'.jpg')
        print ('Picture taken..')
    except Exception as e:
        f = 0
        print ('Picture couldn\'t be taken with exception ' + str(e))

    print

    if(f):
        print ('Scanning image..')
        f = open('foo.jpg','rb')
        qr = PIL.Image.open(f);
        qr.load()

        codes = zbarlight.scan_codes('qrcode',qr)
        if(codes==None):
            #os.remove('qr_codes/qr_'+str(qr_count)+'.jpg')
            print ('No QR code found')
        else:
            print( 'QR code(s):')
            print (codes)
            print (type(codes))
            
            codes1=str(codes)
            print("print=",codes1[21:29])
            key=codes1[21:29]
            
            return key
            
            #res=json.loads(codes1)
            #print(res)

#---------------------------------CAMERA END----------------------------------------------------------





#---------------------------------------DMTR IN START------------------------------------------------
def dmtr_in(p_diff):
    p1=[]
    p_dm=[]
    path_dmtr ="/home/pi/Store Management/trial_store/ledgers_dmtr.xlsx"
    from datetime import datetime
    from openpyxl.styles import Font
    from xlrd import open_workbook
    from openpyxl import load_workbook
    wb = load_workbook(path_dmtr)
    found=0
    IO="IN"
    for data in p_diff:
        p_dm.append(data)
    p1=[]
    pl=p_dm[2]
    #pl=input("Enter PL Number of item: ")
    #print("PL:",pl)

    #EXISTING DATA
    print ("SHEET NAMES: ",wb.sheetnames)
    t1=['Sl.No', 'Date & Time','Specification', 'Quantity', 'Unit', 'Cost', 'Remarks','Type','Total']#list of column  titles for new item
    t=['Sl.No', 'Date & Time', 'Item Name','Specification', 'Quantity', 'Unit', 'Cost', 'Remarks','Type']#list of column  titles
    for sheet in wb.sheetnames:
        if sheet==pl:
            ws=wb.get_sheet_by_name(pl)
            s=ws.max_row-2# variable to store max rows for sl num
            maxr=ws.max_row
            now = datetime.now()# variable to store current date and time
            dt_string = now.strftime("%d/%m/%Y  (%H:%M:%S)")#variable to store current date and time in a format
            
            p=[]#variables to append data
            name=p_dm[3]
            Des=p_dm[10]
            Qty=p_dm[4]
            Unit=p_dm[5]
            Cost=p_dm[6]
            R=p_dm[7]
            tot=0
            if(IO=='IN' or IO=='in'):
                tot=ws.cell(row= maxr,column=9).value+Qty
            elif(IO=='OUT' or IO=='out'):
                tot=ws.cell(row= maxr,column=9).value-Qty
                
            
            sl=s+1
            d=[sl,dt_string,Des,Qty,Unit,Cost,
               R,IO,tot]
            
            print("____________________________________________________________")
            print("\n")
            dm=p1
            p.append(d)
            for data in p:
            
                ws.append(data)
            
            wb.save(path_dmtr)
            
            update_Tot(pl,tot)
            
            
            
            
            
            print("-----------------Data Successfully Added------------------")
            found=1
            break

    #NEW DATA   
    if found==0:
        ws = wb.create_sheet(pl)
        name=p_dm[3]
        ws['A1']="PL:"+pl+"   Description: "+name
        ws['A1'].font=Font(bold=True)
        for cv in range(1,10):
            ws.cell(row=2,column=cv).value=t1[cv-1]
        s=ws.max_row-2# variable to store max rows for sl num
        now = datetime.now()# variable to store current date and time
        dt_string = now.strftime("%d/%m/%Y  (%H:%M:%S)")#variable to store current date and time in a format
        p=[]#variables to append data
        
        
        Des=p_dm[10]
        Qty=p_dm[4]
        Unit=p_dm[5]
        Cost=p_dm[6]
        R=p_dm[7]
        tot=Qty
            
        sl=s+1
        d=[sl,dt_string,Des,Qty,Unit,Cost,R,IO,tot]
        
        print("____________________________________________________________")
        print("\n")
        p.append(d)
        for data in p:
        
            ws.append(data)
        
        wb.save(path_dmtr)
        print("TOT",tot)
        update_Tot(pl,tot)
        found=1 
        print("-----------------Data Successfully Added----------------------")       



    #DMTR
    dmtr= wb.worksheets[0]
    s_dm=dmtr.max_row-1# variable to store max rows for sl num
    sl_dm=s_dm+1

    dm=[sl_dm,dt_string,pl,name,Des,Qty,Unit,Cost,R,IO]
    p1.append(dm)
    for data in p1:

        dmtr.append(data)


    wb.save(path_dmtr)


#--------------------------------------DMTR IN END-----------------------------------------------------


#----------------------------DMTR OUT START---------------------------------------------
def dmtr_out(p_diff):
    p1=[]
    p_dm=[]
    path_dmtr ="/home/pi/Store Management/trial_store/ledgers_dmtr.xlsx"
    from datetime import datetime
    from openpyxl.styles import Font
    from xlrd import open_workbook
    from openpyxl import load_workbook
    wb = load_workbook(path_dmtr)
    found=0
    IO="OUT"
    for data in p_diff:
        p_dm.append(data)
    p1=[]
    pl=p_dm[4]
    #pl=input("Enter PL Number of item: ")
    #print("PL:",pl)

    #EXISTING DATA
    print ("SHEET NAMES: ",wb.sheetnames)
    t1=['Sl.No', 'Date & Time','Specification', 'Quantity', 'Unit', 'Cost', 'Remarks','Type','Total']#list of column  titles for new item
    t=['Sl.No', 'Date & Time', 'Item Name','Specification', 'Quantity', 'Unit', 'Cost', 'Remarks','Type']#list of column  titles
    for sheet in wb.sheetnames:
        if sheet==pl:
            ws=wb.get_sheet_by_name(pl)
            s=ws.max_row-2# variable to store max rows for sl num
            maxr=ws.max_row
            now = datetime.now()# variable to store current date and time
            dt_string = now.strftime("%d/%m/%Y  (%H:%M:%S)")#variable to store current date and time in a format
            
            p=[]#variables to append data
            name=p_dm[5]
            Des=p_dm[10]
            Qty=p_dm[6]
            Unit=p_dm[7]
            Cost="-"
            R=p_dm[11]
            
            tot=0
            
            if(IO=='IN' or IO=='in'):
                tot=ws.cell(row= maxr,column=9).value+Qty
            elif(IO=='OUT' or IO=='out'):
                tot=ws.cell(row= maxr,column=9).value-Qty
                stck1=check_Stock(pl)
                stck=stck1[0]
                print("tot",tot)
                print(type(tot))
                print("stck",stck)
                print(type(stck))
                if(tot<stck):
                    print("UNDERSTOCK")
                
            sl=s+1
            if(tot<0):
                tot=0
            
            d=[sl,dt_string,Des,Qty,Unit,Cost,
               R,IO,tot]
            
            print("____________________________________________________________")
            print("\n")
            dm=p1
            p.append(d)
            for data in p:
            
                ws.append(data)
            
            wb.save(path_dmtr)
            
            update_Tot(pl,tot)
            
            
            
            
            
            print("-----------------Data Successfully Added------------------")
            found=1
            break

    

    #DMTR
    dmtr= wb.worksheets[0]
    s_dm=dmtr.max_row-1# variable to store max rows for sl num
    sl_dm=s_dm+1

    dm=[sl_dm,dt_string,pl,name,Des,Qty,Unit,Cost,R,IO]
    p1.append(dm)
    for data in p1:

        dmtr.append(data)


    wb.save(path_dmtr)


#----------------------------DMTR OUT END---------------------------------------------








#--------------------------------------INTAKE START----------------------------------------------------
def intake():
    path_ch="/home/pi/Store Management/trial_store/check.xlsx"
    e=""
    cnt=0
    stk=[]
    val=" "
    qr_name=" "
    lst=[]
    lst_ch=[]
    qrd=[]
    layout_in= [[sg.Button('NEW'), sg.Button('EXISTING')]]
            
            
            
    
    window=sg.Window('Incoming details', layout_in,size=(300,70),default_element_size=(40,1), grab_anywhere=False)
    event, values = window.read()
   
    if event=='NEW':
        
        
        layout_st=[
                [sg.Button('STOCKED')],
                [sg.Button('NON STOCKED')]
                ]
        
        windowst=sg.Window('STOCK TYPE', layout_st,size=(300,70),default_element_size=(40, 1), grab_anywhere=False)
        event, valuest = windowst.read()
        if event=='NON STOCKED':
            windowst.close()
            layout_ns_new = [
                    #[sg.Text('PL. '), sg.In(key=3)],
                    [sg.Text('Description. '), sg.In(key=4)],
                    [sg.Text('Specification. '), sg.In(key=11)],
                    [sg.Text('Quantity. '), sg.In(key=5)],
                    [sg.Text('Unit. '), sg.In(key=6)],
                    [sg.Text('Cost. '), sg.In(key=7)],
                    [sg.Text('Remarks. '), sg.In(key=8)],
                  
                    [sg.Frame(layout=[
                    #[sg.Checkbox('Checkbox', size=(10,1)),  sg.Checkbox('My second checkbox!', default=True)],
                    [sg.Radio('PO', "RADIO1",), sg.Radio('Letter', "RADIO1",), sg.Radio('Others', "RADIO1",) ]], title='Received Via',title_color='red', relief=sg.RELIEF_SUNKEN, tooltip='Select the mode')],
            
                    [sg.Text('Number. '), sg.In(key=9)],
                    [sg.Text('Minimum Stock. '), sg.In(key=10)],
                    
                    [sg.Button('Save'), sg.Button('Cancel')]
                 ]
        
            window3=sg.Window('Incoming Item Details', layout_ns_new,default_element_size=(40, 1), grab_anywhere=False)
            event, values = window3.read()
            
            '''sg.Popup('Title',
                 'The results of the window.',
                 'The button clicked was "{}"'.format(event),
                 'The values are', values)'''
            if values[0]==True:
                val="PO: "+values[9]
            if values[1]==True:
                val="LT: "+values[9] 
            if values[2]==True:
                val="OT: "+values[9]
            
            if event=='Save':
                values[12]="NON STOCKED"
                wb = load_workbook(path)
                ws = wb.worksheets[0]
                s=ws.max_row# variable to store max rows for sl num
                now = datetime.now()# variable to store current date and time
                sl=s+1
                dt_string = now.strftime("%d/%m/%Y  (%H:%M:%S)")#variable to store current date and time in a format
                
                lst.append(sl)
                lst.append(dt_string)
                values[5]=int(values[5])
                values[10]=int(values[10])
                
                #generating pl_num for new items(NS)
                wb_ns = load_workbook(path_ch)
                ws_ns = wb_ns.worksheets[1]
                s=ws_ns.max_row
                item_num=s
                print(s)
                ns_num=''
                ns='NS'
                zero_app=6-len(str(item_num))
                #print(zero_app)
                for i in range(0,zero_app):
                    ns=ns+'0'
                print(ns)
                ns_num=ns+str(item_num)
                print(ns_num)
                wb_ns.save(path_ch)
                values[3]=ns_num
                
                
                
                
                for i in range(3,9):
                    lst.append(values[i])
                lst.append(val)
                lst.append(values[12])
                print("list=",lst)
                
                
                ws.append(lst)
                
                wb.save(path)
                
                qr_name=lst[2]
                for c in range (2,4):
                    dici={tqr[c]:lst[c]}
                    qrd.append(dici)
                 
                
                img=qrcode.make(qrd)
                qrd=[]
                print(type(img))
                img.save(f"/home/pi/Store Management/trial_store/PL/PL_NUMBER_{qr_name}.png")    
                lst.append(values[11])
                
                dmtr_in(lst)
                
                
                wb_c=load_workbook(path_ch)
                ws1=wb_c.worksheets[1]
                m=ws1.max_row+1
                ws1.cell(row=m,column=1).value=values[3]
                ws1.cell(row=m,column=2).value=values[10]
                ws1.cell(row=m,column=3).value=values[4]
                wb_c.save(path_ch)
                
                
            window3.close()
            sg.popup("DATA SAVED SUCCESSFULLY!")
        
        if event=='STOCKED':
            windowst.close()
            layout_new = [
                    [sg.Text('PL. '), sg.In(key=3)],
                    [sg.Text('Description. '), sg.In(key=4)],
                    [sg.Text('Specification. '), sg.In(key=11)],
                    [sg.Text('Quantity. '), sg.In(key=5)],
                    [sg.Text('Unit. '), sg.In(key=6)],
                    [sg.Text('Cost. '), sg.In(key=7)],
                    [sg.Text('Remarks. '), sg.In(key=8)],
                  
                    [sg.Frame(layout=[
                    #[sg.Checkbox('Checkbox', size=(10,1)),  sg.Checkbox('My second checkbox!', default=True)],
                    [sg.Radio('PO', "RADIO1",), sg.Radio('Letter', "RADIO1",), sg.Radio('Others', "RADIO1",) ]], title='Received Via',title_color='red', relief=sg.RELIEF_SUNKEN, tooltip='Select the mode')],
            
                    [sg.Text('Number. '), sg.In(key=9)],
                    [sg.Text('Minimum Stock. '), sg.In(key=10)],
                    
                    [sg.Button('Save'), sg.Button('Cancel')]
                 ]
        
            window2=sg.Window('Incoming Items', layout_new,default_element_size=(40, 1), grab_anywhere=False)
            event, values = window2.read()
            
            '''sg.Popup('Title',
                 'The results of the window.',
                 'The button clicked was "{}"'.format(event),
                 'The values are', values)'''
            if values[0]==True:
                val="PO: "+values[9]
            if values[1]==True:
                val="LT: "+values[9] 
            if values[2]==True:
                val="OT: "+values[9] 
                
            
            if event=='Save':
                values[12]="STOCKED"
                wb = load_workbook(path)
                ws = wb.worksheets[0]
                s=ws.max_row# variable to store max rows for sl num
                now = datetime.now()# variable to store current date and time
                sl=s+1
                dt_string = now.strftime("%d/%m/%Y  (%H:%M:%S)")#variable to store current date and time in a format
                
                lst.append(sl)
                lst.append(dt_string)
                values[5]=int(values[5])
                values[10]=int(values[10])
                for i in range(3,9):
                    lst.append(values[i])
                lst.append(val)
                lst.append(values[12])
                print("list=",lst)
                
                
                ws.append(lst)
                
                wb.save(path)
                qr_name=lst[2]
                for c in range (2,4):
                    dici={tqr[c]:lst[c]}
                    qrd.append(dici)
                
                
                img=qrcode.make(qrd)
                qrd=[]
                print(type(img))
                img.save(f"/home/pi/Store Management/trial_store/PL/PL_NUMBER_{qr_name}.png")    
                lst.append(values[11])
                
                
                path_ch="/home/pi/Store Management/trial_store/check.xlsx"
                
                wb_c=load_workbook(path_ch)
                ws1=wb_c.worksheets[0]
                m=ws1.max_row+1
                ws1.cell(row=m,column=1).value=values[3]
                ws1.cell(row=m,column=2).value=values[10]
                ws1.cell(row=m,column=3).value=values[4]
                wb_c.save(path_ch)
                dmtr_in(lst)
                
            window2.close()
            sg.popup("DATA SAVED SUCCESSFULLY!")
    
    
    
    
    
    if event=='EXISTING':
        
        layout_st=[
                [sg.Button('STOCKED')],
                [sg.Button('NON STOCKED')]
                ]
        
        windowst=sg.Window('STOCK TYPE', layout_st,size=(300,70),default_element_size=(40, 1), grab_anywhere=False)
        event, valuest = windowst.read()
        
        
        
        
        
        
        if (event=='STOCKED' or event=='NON STOCKED'):
            
            windowst.close()
            if event=='NON STOCKED':
                e='ns'
                #path_ch=path_ns
            elif event=='STOCKED':
                e='s'
                #path_ch="P:/8th SEM/INTERNSHIP/Store Management/trial_store/check.xlsx"
            #window.close()    
            layout_cap=[
                    [sg.Text('PL. '), sg.In(key=3)],[sg.Button('Capture')],
                    [sg.Submit()]
                    ]
            
            layout_exist = [
                    
                    
                    [sg.Text('Specification. '), sg.In(key=11)],
                    [sg.Text('Quantity. '), sg.In(key=5)],
                    [sg.Text('Unit. '), sg.In(key=6)],
                    [sg.Text('Cost. '), sg.In(key=7)],
                    [sg.Text('Remarks. '), sg.In(key=8)],
                  
                    [sg.Frame(layout=[
                    #[sg.Checkbox('Checkbox', size=(10,1)),  sg.Checkbox('My second checkbox!', default=True)],
                    [sg.Radio('PO', "RADIO1",), sg.Radio('Letter', "RADIO1",), sg.Radio('Others', "RADIO1",) ]], title='Received Via',title_color='red', relief=sg.RELIEF_SUNKEN, tooltip='Select the mode')],
            
                    [sg.Text('Number. '), sg.In(key=9)],
                    
                    
                    [sg.Button('Save'), sg.Button('Cancel')]
                 ]
        
            windowcap=sg.Window('Capture PL', layout_cap,default_element_size=(40, 1), grab_anywhere=False)
            event, values1 = windowcap.read()
            '''sg.Popup('Title',
                 'The results of the window.',
                 'The button clicked was "{}"'.format(event),
                 'The values are', values1)'''
            
    
            if event=='Capture':
                print("CAMERA ")
                pl=capture()
                values1[3]=pl
                
            
            #CALL CHECK PL FUNCTION AND GET DESCRIPTION and store it in values[4]
          
            stk=check_Stock(values1[3])
            if stk==None:
                sg.Popup('Record Not Existing')
                windowcap.close()
            elif stk!=None:    
                print(stk)
                windowcap.close()
                window2=sg.Window('Incoming Items', layout_exist,default_element_size=(40, 1), grab_anywhere=False)
                event, values = window2.read()
                values[4]=stk[1]
        
        
                if event=='Save':
                    if e=="s":
                        values[12]="STOCKED"
                    elif e=="ns":
                        values[12]="NON STOCKED"
                    if values[0]==True:
                        val="PO: "+values[9]
                    if values[1]==True:
                        val="LT: "+values[9] 
                    if values[2]==True:
                        val="OT: "+values[9] 
                    wb = load_workbook(path)
                    ws = wb.worksheets[0]
                    s=ws.max_row# variable to store max rows for sl num
                    now = datetime.now()# variable to store current date and time
                    sl=s+1
                    dt_string = now.strftime("%d/%m/%Y  (%H:%M:%S)")#variable to store current date and time in a format
                    
                    lst.append(sl)
                    lst.append(dt_string)
                    lst.append(values1[3])
                    values[5]=int(values[5])
                    for i in range(4,9):
                        lst.append(values[i])
                    lst.append(val)
                    lst.append(values[12])
                    print("list=",lst)
                    '''sg.Popup('Title',
                             'The results of the window.',
                             'The button clicked was "{}"'.format(event),
                             'The values are', values)'''
                    ws.append(lst)
                    wb.save(path)
                    lst.append(values[11])
                    dmtr_in(lst)
                    window2.close()
                    sg.popup("DATA SAVED SUCCESSFULLY!")
    window.close()
    
#---------------------------------------INTAKE END-------------------------------------------------------
    

#---------------------------------------OUT START--------------------------------------------------------

def out():
    tot1=0
    lst=[]
    path ="/home/pi/Store Management/trial_store/test2.xlsx"
    tlm=['Date & Time','Name','Ticket Number','PL', 'Description', 'Quantity', 'Unit', 'Coach Number', 'Coach Type','Type']#list of column  titles
    layout_cap=[
                    [sg.Text('PL. '), sg.In(key=3)],[sg.Button('Capture')],
                    [sg.Submit()]
                    ]
    layout_out = [
                    
                    [sg.Text('Name. '), sg.In(key=13)],
                    [sg.Text('Ticket Number. '), sg.In(key=14)],
                    [sg.Text('Clerk Name. '), sg.In(key=15)],
                    
                    [sg.Text('Specification. '), sg.In(key=11)],
                    [sg.Text('Quantity. '), sg.In(key=5)],
                    [sg.Text('Unit. '), sg.In(key=6)],
                    #[sg.Text('Cost. '), sg.In(key=7)],
                    [sg.Text('Remarks. '), sg.In(key=8)],
                  
                    [sg.Frame(layout=[
                    #[sg.Checkbox('Checkbox', size=(10,1)),  sg.Checkbox('My second checkbox!', default=True)],
                    [sg.Radio('LM', "RADIO1",), sg.Radio('Letter', "RADIO1",), sg.Radio('Others', "RADIO1",) ]], title='Received Via',title_color='red', relief=sg.RELIEF_SUNKEN, tooltip='Select the mode')],
            
                    [sg.Text('Number. '), sg.In(key=9)],
                    [sg.Text('Coach Number. '), sg.In(key=16)],
                    [sg.Text('Coach Type. '), sg.In(key=17)],
                    
                    
                    [sg.Button('Save'), sg.Button('Cancel')]
                 ]
        
    windowcap=sg.Window('Capture PL', layout_cap,default_element_size=(40, 1), grab_anywhere=False)
    event, values1 = windowcap.read()
    
    

    if event=='Capture':
        print("CAMERA ")
        pl=capture()
        values1[3]=pl
    sg.Popup(
         'PL NUMBER', values1[3])
            
            
            
     
    if (values1[3]==None or values1[3]==""):
        sg.Popup('RECORD NOT EXISTING')
        windowcap.close()
    elif (values1[3]!=None and values1[3]!=""):
        tot1=check_Tot1(values1[3]) 
        sg.Popup(
                 'Current stock is:', tot1)
        if (tot1==0):
            sg.Popup("THIS ITEM CANNOT BE ISSUED")
            windowcap.Close()
        elif(tot1!=0):
            window7=sg.Window('Issue Details', layout_out,default_element_size=(40, 1), grab_anywhere=False)
            event, values = window7.read()
            windowcap.Close()
            
            
            if event=='Save':
                wb = load_workbook(path)
                ws = wb.worksheets[1]
                
                
                
                
                lm= wb.worksheets[2]
               
                s=ws.max_row# variable to store max rows for sl num
                sl=s+1
                print("WS row=",s)
                slm=lm.max_row #Second file max row value
                print("LM row=",slm)
                now = datetime.now()# variable to store current date and time
                dt_string = now.strftime("%d/%m/%Y  (%H:%M:%S)")#variable to store current date and time in a format
                p=[]#variables to append data
                d=[]
                
                des_k=[]    
                des_k=check_Stock(values1[3])    
                des_ok=des_k[1]
                values[4]=des_ok
                if values[0]==True:
                    val="LM: "+values[9]
                if values[1]==True:
                    val="LT: "+values[9] 
                if values[2]==True:
                    val="OT: "+values[9]        
                lst.append(sl)
                lst.append(dt_string)
                lst.append(values[13])
                lst.append(values[14])
                lst.append(values1[3])
                lst.append(values[4])
                lst.append(int(values[5]))
                lst.append(values[6])
                lst.append(values[16])
                lst.append(values[17])
                lst.append(values[11])
                lst.append(values[8])
                lst.append(val)
                
                
                
                print("list in lst is=",lst)
                print("\n\n")
                lst1=[]
                
                lst1.append(lst)
               
                print("\n\nlst1 is:",lst1)
                
                
               
                
                
                
                for data in lst1:
            
                    ws.append(data)
                
            
                wb.save(path)
               
                
                
                
                dmtr_out(lst)
                
                
                
                if values[0]==True:
                    cle=values[15]
                    Head=lm.cell(row=slm+2,column=1)#Assign a cell location to Head
                    Head.value="                                                          MATERIAL REQUISITION SLIP"
                    Head1=lm.cell(row=slm+3,column=1)#Assign a cell location to Head
                    Head1.value="Date & Time:"+dt_string#Store the value in Head
                    Head.font=Font(bold=True,color="0000ff" ) #For blue font color="0000ff"
                    Head1.font=Font(bold=True)
                    for cv in range(4,10):
                        lm.cell(row=slm+4,column=cv-3).value=tlm[cv-1]
                        lm.cell(row=slm+4,column=cv-3).font=Font(bold=True)
                    #print("LM row=",slm)
            
                    
                    for data in lst1:
                
                        lm.append(data[4:10])
                    
                        #print("LM row=",slm)
                    end=lm.cell(row=slm+5+1,column=1)
                    end.value="Name:"+lst[2]+"            Ticket Number:"+lst[3]+"                        Clerk:"+cle
                    end.font=Font(bold=True)
                
                            
                
                wb.save(path)
            
           
                    
                window7.Close()
            
    
     
            
        
    



#---------------------------------------OUT END-------------------------------------------------------




#------------------------------------------TOOLS START----------------------------------------
def tools(t,func):
    print("IN TOOLS")
    path ="/home/pi/Store Management/trial_store/Tools.xlsx"
    import openpyxl
    from openpyxl.styles import Font
    from datetime import datetime,timedelta 
    from openpyxl import load_workbook
    from xlrd import open_workbook
    from openpyxl.styles import Font

    #MAIN CONTINUED AT THE BOTTOM









    def inp(tmain):
        wb = load_workbook(path)
        to= wb.worksheets[0]
        t1=tmain
        
        st=to.max_row
        print("row=",st)
        now = datetime.now()# variable to store current date and time
        dt_string = now.strftime("%d/%m/%Y  (%H:%M:%S)")#variable to store current date and time in a format
        p=[]#variables to append data
        d=[]
        Ret=now+timedelta(days=1)
        Ret_str= Ret.strftime("%d/%m/%Y  (%H:%M:%S)")
        Name=t[0]
        TN=t[1]
        ph=t[2]
        clerk=t[3]
        #n=int(input("Enter no of elements: "))
        status=" "
        sl=st
        d=[sl,now,Name,TN,ph,t[4], t[5],int(t[6]),clerk,Ret,status]
        
        
        
        print("____________________________________________________________")
        print("\n")
        p.append(d)
        print("p=",p)
       
           
        
        print("WRITING")
        to.append(d)
            
        wb.save(path)
        sg.popup("TOOL ISSUED")




    def upd(r):
        #path ="P:/8th SEM/INTERNSHIP/Store Management/trial_store/Tools.xlsx"
        
        from openpyxl import load_workbook
        wb=load_workbook(path)
        ws=wb.get_sheet_by_name('Sheet1')
        r1=r+1  
        present=datetime.now() #current time
        prs=present.strftime("%d/%m/%Y  (%H:%M:%S)")
        is1=2#issue date column
        re1=10#return date column
        stat=11#status column
        print("present: ",prs)
        print("RET ",ws.cell(row=r1,column=re1).value)
        if (prs>(ws.cell(row=r1,column=re1).value)):
            ws.cell(row=r1,column=stat).value="NOT RETURNED"
            for i in range (1,12):
                clr=ws.cell(row=ro,column=i)
                clr.font=Font(bold=True,color="ff0000")          
        wb.save(path)



    def Ret_tool(t1):
       t1=t_lst
       from openpyxl import load_workbook
       wb2=load_workbook(path)
       ws_ret=wb2.get_sheet_by_name('Sheet1')
       s=ws_ret.max_row 
       present=datetime.now() #current time
       prs=present.strftime("%d/%m/%Y  (%H:%M:%S)")
       rf=0
       is1=2#issue date column
       re1=10#return date column
       stat=11#status column
       print("present: ",prs)
       Stn=t1[0]
       check_tool=t1[1]
       for ro in range(2,s+1):
           if (((ws_ret.cell(row=ro,column=4).value)==Stn) and (((ws_ret.cell(row=ro,column=stat).value)=="NOT RETURNED") or (ws_ret.cell(row=ro,column=stat).value)==" " )):
               
               if(ws_ret.cell(row=ro,column=6).value==check_tool):
                   print("Record Found")
                   sg.Popup("RECORD FOUND")
                   sg.Popup("TOOL RETURNED")
                   
                   
                   ws_ret.cell(row=ro,column=stat).value="RETURNED"
                   rf=rf+1
                   for i in range (1,12):
                        clr=ws_ret.cell(row=ro,column=i)
                        clr.font=Font(bold=True,color="000000")
       if rf==0:
           print("RECORD NOT FOUND")
           sg.Popup("RECORD NOT FOUND")
                    
       wb2.save(path) 

    #MAIN CONTINUED 
    #while(1):
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    wb=load_workbook(path)
    ws=wb.get_sheet_by_name('Sheet1')
    s=ws.max_row 
    present=datetime.now() #current time
    prs=present.strftime("%d/%m/%Y  (%H:%M:%S)")
    is1=2#issue date column
    re1=10#return date column
    stat=11#status column
    for ro in range(2,s+1):
        if ((present>(ws.cell(row=ro,column=re1).value)) and ((ws.cell(row=ro,column=stat).value)!="RETURNED") ):
            ws.cell(row=ro,column=stat).value="NOT RETURNED"
            for i in range (1,12):
                clr=ws.cell(row=ro,column=i)
                clr.font=Font(bold=True,color="ff0000")    
    wb.save(path)
    
    
    t_lst=[]
    t_lst=t
    #func=int(input("WELCOME TO TOOLS MANAGEMENT\nENTER 1->FOR NEW ENTRY\n 2->FOR RETURN\n 3->EXIT FROM TOOLS:  "))
    fun=func
    if fun==1:
        inp(t_lst)
        
    elif fun==2:
        Ret_tool(t_lst)


#------------------------------------------TOOLS END-----------------------------------------

#----------------------------------------TOOLS GUI---------------------------------------------------
def tool_gui():
    import PySimpleGUI as sg
    import datetime
    lstextra=[]
    lstbase=[]
    func=0
    layout1 = [[sg.Text('TOOLS ')],
              [sg.Button('Issue'), sg.Button('Return')],
              [sg.Exit()]
              ]
    window1 = sg.Window('ORIGINAL').Layout(layout1)
    
    
                # Event Loop
    event, values = window1.Read()
    
    if event == 'Issue':
        window1.close()
        func=1
        layout2 = [
            [sg.Text('Name: '),sg.In(key=0)],
            [sg.Text('Ticket Number: '),sg.In(key=1)],
            [sg.Text('Ph No: '),sg.In(key=2)],
            [sg.Text('Clerk Name: '),sg.In(key=3)],
            [sg.Text('Enter No Of Items: '),sg.InputText(key=4)],
            [sg.Button('Extra Details')]
           ]
        window2 = sg.Window('To issue').Layout(layout2)
        event, values1 = window2.Read()
        for i in range (0,4):
            lstbase.append(values1[i])
        print("LST_BASE=",lstbase)
    
       
        if event == 'Extra Details':
            window2.close()
            layout=[
                   
                    [sg.Text('Tool Borrowed: '),sg.InputText(key=5)],
                    [sg.Text('Description: '),sg.InputText(key=6)],
                    [sg.Text('Quantity: '),sg.InputText(key=7)],
                    [sg.Button('Next')]
                   ]
            window = sg.Window('To enter extra details').Layout(layout)
            #event, values = window.Read()
            
            
            i=0
            while i<int(values1[4]):                             # The Event Loop
                lstextra=[]
                for ap in range (0,4):
                    lstextra.append(lstbase[ap])    
                
                event, values = window.read()
                if event == 'Next':
                    for w in range(5,8):
                        lstextra.append(values[w])
                        window[w].update('')
                    #print(event, values)
                    i=i+1  
                    
                
                    print("base",lstbase)
                print("Final",lstextra)
    
                tools(lstextra,func)
            
                      
            #if event=='Submit':
             #   tools(lstextra)
            window.close()
    
            
        #if event == 'Submit':
         #   window.Close()
            
          #  break
                        
        window2.Close()
        
        
    elif event == 'Return':
        window1.Close()
        func=2
        layout3 = [
            [sg.Text('Ticket Number: '),sg.InputText(key=8)],
            [sg.Text('Tool: '),sg.InputText(key=9)],
            [sg.Button('Submit')]
           ]
    
        window3 = sg.Window('To return').Layout(layout3)
        event, values = window3.Read()
    
        for i in range (8,10):
            lstbase.append(values[i])
        print("LST_BASE=",lstbase)
    
        if event == 'Submit':
            tools(lstbase,func)
            window3.Close()

    elif event=='Exit':
        window1.close()

#----------------------------------------TOOLS GUI END-----------------------------------------------














  
    
#----------------------------------------MAIN START-------------------------------------------------------    
sg.theme('SystemDefault1')  # please make your windows colorful size=(15, 1)),background_color=''

sg.SetOptions(element_padding=(0, 0))




layout = [
                        
                        [sg.Image("/home/pi/Store Management/trial_store/south.PNG")],
                        [sg.Text('\n')],
                        [sg.Text('\n')],
                        [sg.Text('\n')], 
                        
                        [sg.T(' '  * 160),sg.Text('STORE MANAGEMENT',text_color='blue',font=('Helvetica', 25))],
                        [sg.Text('\n')],
                        
                        [sg.T(' '  * 80),sg.Button('IN',button_color=('white', 'blue'),size=(10, 2), font=("Helvetica", 15)),
                        sg.T(' '  * 10), sg.Button('OUT',button_color=('white', 'blue'),size=(10, 2), font=("Helvetica", 15)),
                         
                        sg.T(' '  * 10),sg.Button('CHECK STOCK',button_color=('white', 'blue'),size=(10, 2), font=("Helvetica", 15)),
                        sg.T(' '  * 10), sg.Button('SEARCH',button_color=('white', 'blue'),size=(10, 2), font=("Helvetica", 15)),
                         
                        sg.T(' '  * 10),sg.Button('TOOLS',button_color=('white', 'blue'),size=(10, 2), font=("Helvetica", 15)),
                        sg.T(' '  * 10), sg.Button('UNDER STOCK',button_color=('white', 'blue'),size=(10, 2), font=("Helvetica", 15))],
                        
                        
                        [sg.Text('\n')],
                        [sg.Text('\n')],
                        [sg.Text('\n')],
                        [sg.Text('\n')],
                        [sg.Text('\n')],
                        [sg.Text('\n')],
                        [sg.T(' '  * 180),sg.Exit(button_color=('white', 'red'),size=(8, 1),font=("Helvetica", 12))],
			#[sg.T(' '  * 160),sg.Submit(button_color=('white', 'red')),sg.T(' '  * 15), sg.Cancel(button_color=('white', 'red'))],
                        [sg.Text('\n')]
                        
			]



window_main = sg.Window('Simple data entry window', layout, location=(0,0), finalize=True)
window_main.Maximize()

#sg.Popup(event, values, values['-NAME-'], values['-ADDRESS-'], values['-PHONE-'])


while True:             # Event Loop
    event, values = window_main.Read()
    if event in (None, 'Exit'):
        break
    if event == 'IN':
        intake()
    elif event == 'OUT':
        out()
    elif event == 'CHECK STOCK':
        check_Tot()
    elif event == 'SEARCH':
        search()
    elif event == 'TOOLS':
        tool_gui()
    elif event == 'UNDER STOCK':
        min_check()
window_main.Close()   
    
 #---------------------------------------------MAIN END-------------------------------------------------   
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
