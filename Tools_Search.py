# -*- coding: utf-8 -*-
"""
Created on Wed Apr 29 15:58:53 2020

@author: Rohan Roy
"""

path ="P:/8th SEM/INTERNSHIP/Store Management/trial_store/Tools.xlsx"
import openpyxl
from openpyxl.styles import Font
from datetime import datetime,timedelta 
from openpyxl import load_workbook
from xlrd import open_workbook
from openpyxl.styles import Font

#MAIN CONTINUED AT THE BOTTOM









def inp():
    wb = load_workbook(path)
    to= wb.worksheets[0]
    
    
    st=to.max_row
    print("row=",st)
    now = datetime.now()# variable to store current date and time
    dt_string = now.strftime("%d/%m/%Y  (%H:%M:%S)")#variable to store current date and time in a format
    p=[]#variables to append data
    d=[]
    Ret=now+timedelta(days=1)
    Ret_str= Ret.strftime("%d/%m/%Y  (%H:%M:%S)")
    Name=input("Enter Name: ")
    TN=input("Enter Ticket Number: ")
    ph=input("Enter Phone Number: ")
    clerk=input("Input Clerk Name: ")
    n=int(input("Enter no of elements: "))
    status=" "
    for a in range(0,n):
        print("enter the details of item",a+1)
        sl=st+a
        d=[sl,now,Name,TN,ph,input("Tool borrowed: "), input("Enter Description: "),int(input("Enter Quantity: ")),clerk,Ret,status]
        
        
        
        print("____________________________________________________________")
        print("\n")
        p.append(d)
        print("p=",p)
       
       
    for data in p:
    
        to.append(data)
        
    wb.save(path)




def upd(r):
    #path ="P:/8th SEM/INTERNSHIP/Store Management/trial_store/Tools.xlsx"
    
    from openpyxl import load_workbook
    wb=load_workbook(path)
    ws=wb.get_sheet_by_name('Sheet1')
    r1=r+1  
    present=datetime.now() #current time
    prs=present.strftime("%d/%m/%Y  (%H:%M:%S)")
    is1=2#issue date column
    re1=10#return date column
    stat=11#status column
    print("present: ",prs)
    print("RET ",ws.cell(row=r1,column=re1).value)
    if (prs>(ws.cell(row=r1,column=re1).value)):
        ws.cell(row=r1,column=stat).value="NOT RETURNED"
        for i in range (1,12):
            clr=ws.cell(row=ro,column=i)
            clr.font=Font(bold=True,color="ff0000")          
    wb.save(path)



def Ret_tool():
   from openpyxl import load_workbook
   wb2=load_workbook(path)
   ws_ret=wb2.get_sheet_by_name('Sheet1')
   s=ws_ret.max_row 
   present=datetime.now() #current time
   prs=present.strftime("%d/%m/%Y  (%H:%M:%S)")
   rf=0
   is1=2#issue date column
   re1=10#return date column
   stat=11#status column
   print("present: ",prs)
   Stn=input("Enter Ticket Number: ")
   check_tool=input("Enter Tool Name: ")
   for ro in range(2,s+1):
       if (((ws_ret.cell(row=ro,column=4).value)==Stn) and (((ws_ret.cell(row=ro,column=stat).value)=="NOT RETURNED") or (ws_ret.cell(row=ro,column=stat).value)==" " )):
           
           if(ws_ret.cell(row=ro,column=6).value==check_tool):
               print("Record Found")
        
               ws_ret.cell(row=ro,column=stat).value="RETURNED"
               rf=rf+1
               for i in range (1,12):
                    clr=ws_ret.cell(row=ro,column=i)
                    clr.font=Font(bold=True,color="000000")
   if rf==0:
       print("RECORD NOT FOUND")
    
                
   wb2.save(path) 

#MAIN CONTINUED 
while(1):
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    wb=load_workbook(path)
    ws=wb.get_sheet_by_name('Sheet1')
    s=ws.max_row 
    present=datetime.now() #current time
    prs=present.strftime("%d/%m/%Y  (%H:%M:%S)")
    is1=2#issue date column
    re1=10#return date column
    stat=11#status column
    for ro in range(2,s+1):
        if ((present>(ws.cell(row=ro,column=re1).value)) and ((ws.cell(row=ro,column=stat).value)!="RETURNED") ):
            ws.cell(row=ro,column=stat).value="NOT RETURNED"
            for i in range (1,12):
                clr=ws.cell(row=ro,column=i)
                clr.font=Font(bold=True,color="ff0000")    
    wb.save(path)

    
    
    
    func=int(input("WELCOME TO TOOLS MANAGEMENT\nENTER 1->FOR NEW ENTRY 2->FOR RETURN:  "))
    if func==1:
        inp()
    elif func==2:
        Ret_tool()
    elif func==3:
        break
    else:
        print("ERROR")
    
        

#inp()        
#search_tool()
#Ret_tool()